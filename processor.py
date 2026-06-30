"""旅行命令復命書 Excel から動静表 Excel を生成するコア処理."""

from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from typing import BinaryIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# 出力1行を構成する4項目とその既定の左から右への並び順
DEFAULT_ORDER = ["name", "purpose", "dest", "time"]
VALID_KEYS = set(DEFAULT_ORDER)

# 各項目をどう装飾して表示するか（並び順に関わらず項目ごとに固定）
_RENDERERS = {
    "name": lambda v: f"{v}：" if v else "",
    "purpose": lambda v: v,
    "dest": lambda v: f"（{v}）" if v else "",
    "time": lambda v: v,
}


def _render_line(fields: dict[str, str], order: list[str]) -> str:
    return "".join(_RENDERERS[key](fields[key]) for key in order)


def _extract_entries(workbook: openpyxl.Workbook) -> dict[tuple[int, int, int], list[dict[str, str]]]:
    by_date: dict[tuple[int, int, int], list[dict[str, str]]] = defaultdict(list)

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        name = ws["W3"].value
        reiwa_y = ws["J4"].value
        month = ws["M4"].value
        day = ws["P4"].value
        start_h = ws["S4"].value
        start_m = ws["V4"].value
        end_h = ws["S5"].value
        end_m = ws["V5"].value
        dest = ws["F7"].value
        purpose = ws["F8"].value

        if not name or reiwa_y is None or month is None or day is None:
            continue

        date_key = (int(reiwa_y), int(month), int(day))
        time_str = (
            f"{int(start_h or 0):02d}:{int(start_m or 0):02d}"
            f"~{int(end_h or 0):02d}:{int(end_m or 0):02d}"
        )
        fields = {
            "name": str(name).strip(),
            "purpose": str(purpose).strip() if purpose else "",
            "dest": str(dest).strip() if dest else "",
            "time": time_str,
        }
        by_date[date_key].append(fields)

    return by_date


def _build_workbook(
    by_date: dict[tuple[int, int, int], list[dict[str, str]]], order: list[str]
) -> openpyxl.Workbook:
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "動静表"

    out_ws["A1"] = "日付"
    out_ws["B1"] = "予定"
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in (out_ws["A1"], out_ws["B1"]):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row = 2
    for date_key in sorted(by_date.keys()):
        reiwa_y, month, day = date_key
        date_str = f"令和{reiwa_y}年{month}月{day}日"
        combined = "\n".join(_render_line(fields, order) for fields in by_date[date_key])

        out_ws.cell(row=row, column=1, value=date_str)
        out_ws.cell(row=row, column=2, value=combined)
        out_ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        out_ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        out_ws.cell(row=row, column=1).border = border
        out_ws.cell(row=row, column=2).border = border

        line_count = len(by_date[date_key])
        out_ws.row_dimensions[row].height = max(20, line_count * 18)
        row += 1

    out_ws.column_dimensions["A"].width = 18
    out_ws.column_dimensions["B"].width = 90
    out_ws.freeze_panes = "A2"

    return out_wb


def _format_output_name(by_date: dict[tuple[int, int, int], list[dict[str, str]]]) -> str:
    """期間に応じてファイル名を組み立てる.

    - 単日             : 動静表_令和8年6月2日.xlsx
    - 同年同月         : 動静表_令和8年6月2日～30日.xlsx
    - 同年異月         : 動静表_令和8年6月2日～7月15日.xlsx
    - 異年             : 動静表_令和8年6月2日～令和9年1月15日.xlsx
    """
    if not by_date:
        return "動静表.xlsx"

    dates = sorted(by_date.keys())
    s_y, s_m, s_d = dates[0]
    e_y, e_m, e_d = dates[-1]

    start = f"令和{s_y}年{s_m}月{s_d}日"
    if (s_y, s_m, s_d) == (e_y, e_m, e_d):
        return f"動静表_{start}.xlsx"
    if s_y == e_y and s_m == e_m:
        return f"動静表_{start}～{e_d}日.xlsx"
    if s_y == e_y:
        return f"動静表_{start}～{e_m}月{e_d}日.xlsx"
    return f"動静表_{start}～令和{e_y}年{e_m}月{e_d}日.xlsx"


def process_excel(
    file: BinaryIO | str, order: list[str] | None = None
) -> tuple[bytes, str, int, int]:
    """入力 Excel を処理し、(出力バイト列, 出力ファイル名, 日数, 件数) を返す.

    order: 1行内の項目の左から右への並び順。"name", "purpose", "dest", "time" の
    4つをすべて含む必要がある。省略時は DEFAULT_ORDER を使う。
    """
    if order is None:
        order = DEFAULT_ORDER
    elif set(order) != VALID_KEYS:
        raise ValueError(
            f"order には {sorted(VALID_KEYS)} を重複なくすべて含めてください: {order}"
        )

    wb = openpyxl.load_workbook(file, data_only=True)
    by_date = _extract_entries(wb)
    out_wb = _build_workbook(by_date, order)
    output_name = _format_output_name(by_date)

    buf = BytesIO()
    out_wb.save(buf)
    buf.seek(0)

    total_entries = sum(len(v) for v in by_date.values())
    return buf.getvalue(), output_name, len(by_date), total_entries
